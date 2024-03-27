from .models import Ventas
from rest_framework.views import APIView
from rest_framework import status, permissions
from .serializers import ventasSerializer
from rest_framework.response import Response

from openpyxl import load_workbook

class uploadManual(APIView):
    permission_classes = (permissions.AllowAny, )
    
    def post(self, request):
        serializer = ventasSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
class import_excel(APIView):
    permission_classes = (permissions.AllowAny, )
    
    def post(self, request):
        try:
            
            excel_file = request.FILES['file']
            wb = load_workbook(excel_file)
            ws = wb.active
            error_messages = []
            
            # Extraer los encabezados y los valores
            headers = [cell.value for cell in ws[1]]            
        
            for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                row_data = dict(zip(headers,row))
                print(row_data)
                if any(value is not None for value in row[:len(headers)]):
                    try:
                        
                        data = Ventas(
                            nombre_doc=row_data['Nombre Doc'],                      \
                            num_doc=row_data['Número del Documento'],               \
                            cod_cliente=row_data['RUT Cliente'],                    \
                            nom_cliente=row_data['Nombre del Cliente'],             \
                            fecha=row_data['Fecha'],                                \
                            fecha_venc=row_data['Fecha Vencimiento'],               \
                            desc_producto=row_data['Desc. Producto'],               \
                            total_detalle=row_data['Total Detalle'],                \
                            analisis_cn=(row_data['Análisis C. Negocio']),          \
                            comentario=row_data['Comentario Ítem'],                 \
                            linea=row_data['Línea'],                                \
                            empresa=row_data['Empresa'],                            \
                            precio_unit=row_data['Precio Unitario'],                \
                            total_neto=row_data['Total Neto Linea'],                        \
                            es_venta=row_data['Es venta']                                   \
                        )
                        print(data)
                        data.full_clean()
                        data.save()
                    except KeyError as err:
                        error_messages.append(f"Error en la fila {row_number}: {err}")
                else:
                    continue
            if error_messages:
                return Response({'error': error_messages}, status=status.HTTP_400_BAD_REQUEST)
            else:
                return Response({'mensaje': 'Datos cargados correctamente'}, status=status.HTTP_201_CREATED)
        except KeyError:
            return Response({"error": "No se proporcionó el archivo correctamente"}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as err:
            return Response({"error": str(err)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            

            
            